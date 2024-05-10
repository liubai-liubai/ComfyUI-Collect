import json
import time
from openpyxl import load_workbook
import textwrap

import google.generativeai as genai

from IPython.display import display
from IPython.display import Markdown


def to_markdown(text):
  text = text.replace('•', '  *')
  return Markdown(textwrap.indent(text, '> ', predicate=lambda _: True))

# 读取config.json文件中的token
with open('config.json') as f:
    config = json.load(f)
    gemini_key = config['gemini_key']

#获取gemini-key
genai.configure(api_key=gemini_key)
model = genai.GenerativeModel('gemini-pro')



# 加载 Excel 文件
wb = load_workbook('github_repos.xlsx')
ws = wb.active

user_input = ""

# 初始化计数器
counter = 0
_system_instructions = f"当我发送给你一个网址后，如果没访问成功则回复访问失败，成功了则总结这个网址访问后的要点，要点不超过100字且标题为概述，然后详细分析网址中项目的每个功能，详细分析不限字数。其他废话不要多说。"
    
model._system_instructions = _system_instructions

# 遍历除第一行之外的每一行
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    if row[4].value is not None:
        continue
    url = row[3].value  # 获取第4列的值（即网址）
    user_input = str(url)
    print(user_input)
 

    response = model.generate_content(user_input)
   
    result = to_markdown(response.text)  

    print(result)

     # 将响应写入到第5列
    ws.cell(row=i, column=5, value=result)

     # 增加计数器
    counter += 1

     # 如果计数器达到10，保存 Excel 文件并重置计数器
    if counter == 5:
        wb.save('github_repos.xlsx')
        counter = 0

    time.sleep(0.3)
# 保存 Excel 文件
wb.save('github_repos.xlsx')