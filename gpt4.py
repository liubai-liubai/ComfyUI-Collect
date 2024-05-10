import json
import time
from openai import OpenAI
from openpyxl import load_workbook

# 读取config.json文件中的token
with open('config.json') as f:
    config = json.load(f)
    gpt_key = config['gpt_key']

# 替换为您的 Moonshot AI API 密钥
client = OpenAI(
    api_key=f"{gpt_key}"
)

completion = client.chat.completions.create(
  model="gpt-3.5-turbo",
  messages=[
    {"role": "system", "content": "You are a poetic assistant, skilled in explaining complex programming concepts with creative flair."},
    {"role": "user", "content": "Compose a poem that explains the concept of recursion in programming."}
  ]
)

print(completion.choices[0].message)


# # 准备对话历史和用户输入
# history = [
#     {
#         "role": "system",
#         "content": "发送给你连接后，对这个连接进行访问，如果没访问成功则输出访问失败，如果访问成功，则总结这个网址内项目的要点，要点不超过100字且标题为概述，然后详细分析网址中项目的主要的作用，旨在让我明白这个项目是干啥用的，有哪些功能，这个不限字数。其他废话不要多说。"
#     }
# ]

# # 加载 Excel 文件
# wb = load_workbook('github_repos.xlsx')
# ws = wb.active

# user_input = ""

# # 初始化计数器
# counter = 0

# # 遍历除第一行之外的每一行
# for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
#     if row[4].value is not None:
#         continue
#     url = row[3].value  # 获取第4列的值（即网址）

#     # 准备对话历史和用户输入
#     history = [
#         {
#             "role": "system",
#             "content": f"访问发送给你连接后{url}，如果没访问成功则回复访问失败，成功了则总结这个网址的要点，要点不超过100字且标题为概述，然后详细分析网址中项目的每个功能，这个不限字数。其他废话不要多说。"
#         }
#     ]

#     user_input = f"你好，我想了解更多关于这个网站的详细信息。{url}"
#     print(user_input)
#     # 发送请求
#     completion = client.chat.completions.create(
#         model="moonshot-v1-32k",
#         messages=history + [{"role": "user", "content": user_input}],
#         temperature=0.3,
#     )

#     # 获取响应
#     response = completion.choices[0].message.content
#     print(response)
#      # 将响应写入到第5列
#     ws.cell(row=i, column=5, value=response)

#      # 增加计数器
#     counter += 1

#      # 如果计数器达到10，保存 Excel 文件并重置计数器
#     if counter == 5:
#         wb.save('github_repos.xlsx')
#         counter = 0

#     time.sleep(0.3)
# # 保存 Excel 文件
# wb.save('github_repos.xlsx')